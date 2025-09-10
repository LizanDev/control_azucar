import os
import sys

# Configurar variables de entorno ANTES de cualquier otra importación
from dotenv import load_dotenv

def cargar_env_ejecutable():
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)
        env_path = os.path.join(exe_dir, '.env')
        if os.path.exists(env_path):
            load_dotenv(env_path)
        else:
            # Configuración por defecto sin exponer claves sensibles
            os.environ['ABACUS_API_KEY'] = '[CLAVE_API_REMOVIDA]'
            os.environ['ABACUS_API_URL'] = '[URL_API_REMOVIDA]'
            os.environ['APP_NAME'] = 'Control de Azúcar y Alimentación'
            os.environ['DATA_FILE'] = 'control_alimentacion.json'
    else:
        load_dotenv()

cargar_env_ejecutable()

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
from PIL.ExifTags import TAGS
import json
import requests
import base64
from datetime import datetime
import exifread

class AbacusAIClient:
    """Cliente para interactuar con la API de Abacus.AI"""

    def __init__(self):
        self.api_key = os.getenv('ABACUS_API_KEY')
        self.api_url = os.getenv('ABACUS_API_URL')

        if not self.api_key:
            if getattr(sys, 'frozen', False):
                raise ValueError("Error de configuración en el ejecutable. Contacta al desarrollador.")
            else:
                raise ValueError("ABACUS_API_KEY no encontrada en el archivo .env")
        if not self.api_url:
            if getattr(sys, 'frozen', False):
                raise ValueError("Error de configuración en el ejecutable. Contacta al desarrollador.")
            else:
                raise ValueError("ABACUS_API_URL no encontrada en el archivo .env")

    def encode_image_to_base64(self, image_path):
        """Convertir imagen a base64 para enviar a la API"""
        try:
            with open(image_path, "rb") as image_file:
                return base64.b64encode(image_file.read()).decode('utf-8')
        except Exception as e:
            raise Exception(f"Error al codificar imagen: {str(e)}")

    def identificar_alimentos(self, image_path):
        """
        Identificar alimentos en una imagen usando Abacus.AI
        """
        try:
            # Codificar imagen a base64
            base64_image = self.encode_image_to_base64(image_path)

            # Preparar el payload para la API
            headers = {
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json"
            }

            payload = {
                "model": "gpt-4o",  # Modelo con capacidades de visión
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": """Analiza esta imagen de alimentos y devuelve ÚNICAMENTE una lista de los alimentos que puedes identificar, separados por comas. 

Ejemplo de respuesta: "Manzana, Pan integral, Huevos revueltos, Café, Yogur"

Solo menciona alimentos específicos que puedas ver claramente en la imagen. No agregues explicaciones adicionales."""
                            },
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpeg;base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                "max_tokens": 300,
                "temperature": 0.1
            }

            # Realizar petición a la API
            response = requests.post(self.api_url, headers=headers, json=payload, timeout=30)

            if response.status_code == 200:
                result = response.json()

                # Extraer la respuesta del modelo
                if 'choices' in result and len(result['choices']) > 0:
                    content = result['choices'][0]['message']['content'].strip()

                    # Procesar la respuesta para extraer lista de alimentos
                    alimentos = [alimento.strip() for alimento in content.split(',')]
                    alimentos = [alimento for alimento in alimentos if alimento]  # Filtrar vacíos

                    return alimentos
                else:
                    raise Exception("Respuesta inesperada de la API")

            else:
                error_msg = f"Error {response.status_code}: {response.text}"
                raise Exception(error_msg)

        except requests.exceptions.Timeout:
            raise Exception("Timeout: La API tardó demasiado en responder")
        except requests.exceptions.ConnectionError:
            raise Exception("Error de conexión: No se pudo conectar con la API")
        except Exception as e:
            raise Exception(f"Error al identificar alimentos: {str(e)}")

class ControlAzucarApp:
    def __init__(self, root):
        self.root = root
        self.root.title(os.getenv('APP_NAME', 'Control de Azúcar y Alimentación'))
        self.root.geometry("850x700")

        # Inicializar cliente de Abacus.AI
        try:
            self.ai_client = AbacusAIClient()
        except Exception as e:
            messagebox.showerror("Error de Configuración", 
                               f"Error al inicializar Abacus.AI: {str(e)}\n\n"
                               "Verifica tu archivo .env")
            self.ai_client = None

        # Datos de la aplicación
        self.datos_file = os.getenv('DATA_FILE', 'control_alimentacion.json')
        self.cargar_datos()

        self.crear_interfaz()

    def extraer_fecha_hora_exif(self, ruta_imagen):
        """Extraer fecha y hora de los metadatos EXIF de una imagen"""
        try:
            # Método 1: Usando PIL
            with Image.open(ruta_imagen) as img:
                exifdata = img.getexif()
                
                # Buscar fecha en diferentes campos EXIF
                fecha_campos = [36867, 36868, 306]  # DateTime, DateTimeOriginal, DateTimeDigitized
                
                for campo in fecha_campos:
                    if campo in exifdata:
                        fecha_str = exifdata[campo]
                        try:
                            # Formato EXIF: "YYYY:MM:DD HH:MM:SS"
                            fecha_objeto = datetime.strptime(fecha_str, "%Y:%m:%d %H:%M:%S")
                            return {
                                'fecha': fecha_objeto.strftime("%Y-%m-%d"),
                                'hora': fecha_objeto.strftime("%H:%M"),
                                'datetime': fecha_objeto,
                                'fuente': 'EXIF'
                            }
                        except ValueError:
                            continue
                            
            # Método 2: Usando exifread como respaldo
            with open(ruta_imagen, 'rb') as f:
                tags = exifread.process_file(f)
                
                fecha_tags = ['EXIF DateTimeOriginal', 'EXIF DateTime', 'Image DateTime']
                
                for tag_name in fecha_tags:
                    if tag_name in tags:
                        fecha_str = str(tags[tag_name])
                        try:
                            fecha_objeto = datetime.strptime(fecha_str, "%Y:%m:%d %H:%M:%S")
                            return {
                                'fecha': fecha_objeto.strftime("%Y-%m-%d"),
                                'hora': fecha_objeto.strftime("%H:%M"),
                                'datetime': fecha_objeto,
                                'fuente': 'EXIF'
                            }
                        except ValueError:
                            continue
                            
        except Exception as e:
            print(f"Error al extraer EXIF: {e}")
            
        # Si no se pudo extraer de EXIF, usar fecha/hora actual
        ahora = datetime.now()
        return {
            'fecha': ahora.strftime("%Y-%m-%d"),
            'hora': ahora.strftime("%H:%M"),
            'datetime': ahora,
            'fuente': 'Actual'
        }

    def cargar_datos(self):
        """Cargar datos existentes o crear estructura inicial"""
        try:
            with open(self.datos_file, 'r', encoding='utf-8') as f:
                self.datos = json.load(f)
        except FileNotFoundError:
            self.datos = {
                "registros": [],
                "configuracion": {
                    "franjas_horarias": {
                        "desayuno": {"inicio": "06:00", "fin": "10:00"},
                        "almuerzo": {"inicio": "10:01", "fin": "13:00"},
                        "comida": {"inicio": "13:01", "fin": "16:00"},
                        "merienda": {"inicio": "16:01", "fin": "19:00"},
                        "cena": {"inicio": "19:01", "fin": "23:59"}
                    }
                }
            }

    def guardar_datos(self):
        """Guardar datos en archivo JSON"""
        with open(self.datos_file, 'w', encoding='utf-8') as f:
            json.dump(self.datos, f, ensure_ascii=False, indent=2)

    # NOTA: Esta función ya no se usa, ahora se usan nombres personalizados
    # def determinar_comida_por_hora(self, hora_str):
    #     """Determinar tipo de comida según la hora (OBSOLETO)"""
    #     hora = datetime.strptime(hora_str, "%H:%M").time()
    #     franjas = self.datos["configuracion"]["franjas_horarias"]
    #
    #     for comida, franja in franjas.items():
    #         inicio = datetime.strptime(franja["inicio"], "%H:%M").time()
    #         fin = datetime.strptime(franja["fin"], "%H:%M").time()
    #
    #         if inicio <= hora <= fin:
    #             return comida
    #
    #     return "otro"

    def crear_interfaz(self):
        """Crear la interfaz gráfica"""
        # Frame principal con scroll
        canvas = tk.Canvas(self.root)
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        main_frame = ttk.Frame(scrollable_frame, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Título
        titulo = ttk.Label(main_frame, text="🍎 Control de Azúcar y Alimentación con IA", 
                          font=("Arial", 18, "bold"))
        titulo.pack(pady=(0, 25))

        # Estado de la API
        if self.ai_client:
            estado_api = ttk.Label(main_frame, text="✅ Conectado con Abacus.AI", 
                                 foreground="green", font=("Arial", 10))
        else:
            estado_api = ttk.Label(main_frame, text="❌ Error de conexión con Abacus.AI", 
                                 foreground="red", font=("Arial", 10))
        estado_api.pack(pady=(0, 15))

        # Sección de registro de azúcar
        azucar_frame = ttk.LabelFrame(main_frame, text="📊 Registro de Comida y Azúcar", padding="10")
        azucar_frame.pack(fill=tk.X, pady=(0, 15))

        # Campo para nombre de la comida
        nombre_comida_frame = ttk.Frame(azucar_frame)
        nombre_comida_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(nombre_comida_frame, text="🍽️ Nombre de la comida:").pack(side=tk.LEFT)
        self.nombre_comida_var = tk.StringVar()
        nombre_comida_entry = ttk.Entry(nombre_comida_frame, textvariable=self.nombre_comida_var, width=25)
        nombre_comida_entry.pack(side=tk.LEFT, padx=(10, 5))
        
        # Botón para sugerencias
        ttk.Button(nombre_comida_frame, text="💡", width=3,
                  command=self.mostrar_sugerencias_comida).pack(side=tk.LEFT, padx=(0, 5))
        
        # Etiqueta de ayuda
        ttk.Label(nombre_comida_frame, text="(ej: Desayuno, Almuerzo, Merienda)", 
                 foreground="gray", font=("Arial", 8)).pack(side=tk.LEFT, padx=(5, 0))

        azucar_input_frame = ttk.Frame(azucar_frame)
        azucar_input_frame.pack(fill=tk.X, pady=(0, 10))

        # Campo para azúcar antes y después de comer en una sola fila
        azucar_frame = ttk.Frame(azucar_input_frame)
        azucar_frame.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(azucar_frame, text="📉 Azúcar ANTES de comer:").pack(side=tk.LEFT)
        self.azucar_antes_var = tk.StringVar()
        azucar_antes_entry = ttk.Entry(azucar_frame, textvariable=self.azucar_antes_var, width=15)
        azucar_antes_entry.pack(side=tk.LEFT, padx=(10, 5))
        ttk.Label(azucar_frame, text="mg/dL").pack(side=tk.LEFT, padx=(5, 20))

        ttk.Label(azucar_frame, text="📈 Azúcar DESPUÉS de comer:").pack(side=tk.LEFT)
        self.azucar_despues_var = tk.StringVar()
        azucar_despues_entry = ttk.Entry(azucar_frame, textvariable=self.azucar_despues_var, width=15)
        azucar_despues_entry.pack(side=tk.LEFT, padx=(10, 5))
        ttk.Label(azucar_frame, text="mg/dL").pack(side=tk.LEFT)

        # Nota informativa
        nota_frame = ttk.Frame(azucar_input_frame)
        nota_frame.pack(fill=tk.X)
        ttk.Label(nota_frame, text="💡 Tip: Puedes llenar solo uno de los campos si no tienes ambos valores",
                 foreground="gray", font=("Arial", 8), wraplength=400).pack()

        # Sección de foto de alimentos
        foto_frame = ttk.LabelFrame(main_frame, text="📸 Análisis de Alimentos con IA", padding="10")
        foto_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        # Botón para seleccionar foto
        ttk.Button(foto_frame, text="📁 Seleccionar Foto de Alimentos", 
                  command=self.seleccionar_foto).pack(pady=(0, 10))

        # Label para mostrar información de metadatos
        self.info_metadata_label = ttk.Label(foto_frame, text="", 
                                           font=("Arial", 9, "italic"))
        # No hacer pack aquí, se mostrará solo cuando se seleccione una foto

        # Frame para mostrar la imagen
        self.imagen_frame = ttk.Frame(foto_frame)
        self.imagen_frame.pack(pady=10)

        self.imagen_label = ttk.Label(self.imagen_frame, text="📷 No hay imagen seleccionada", 
                                    font=("Arial", 10), foreground="gray")
        self.imagen_label.pack()

        # Botón de análisis
        self.btn_analizar = ttk.Button(foto_frame, text="🤖 Analizar con IA", 
                                     command=self.analizar_foto, state="disabled")
        self.btn_analizar.pack(pady=10)

        # Progress bar para análisis
        self.progress = ttk.Progressbar(foto_frame, mode='indeterminate')
        self.progress.pack(fill=tk.X, pady=(0, 10))

        # Lista de alimentos identificados
        alimentos_frame = ttk.LabelFrame(foto_frame, text="🥗 Alimentos Identificados", padding="10")
        alimentos_frame.pack(fill=tk.BOTH, expand=True)

        self.alimentos_listbox = tk.Listbox(alimentos_frame, height=8, font=("Arial", 10))
        self.alimentos_listbox.pack(fill=tk.BOTH, expand=True)

        # Botones de acción
        botones_frame = ttk.Frame(main_frame)
        botones_frame.pack(fill=tk.X, pady=(10, 0), anchor="e")
        ttk.Button(botones_frame, text="Guardar Registro", command=self.guardar_registro).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(botones_frame, text="Mostrar Historial", command=self.mostrar_historial).pack(side=tk.RIGHT, padx=(5, 0))

        # Configurar scroll
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Variables para la foto
        self.foto_path = None
        self.alimentos_detectados = []

    def seleccionar_foto(self):
        """Seleccionar foto de alimentos y extraer metadatos"""
        self.foto_path = filedialog.askopenfilename(
            title="Seleccionar foto de alimentos",
            filetypes=[
                ("Imágenes", "*.jpg *.jpeg *.png *.bmp *.gif"),
                ("JPEG", "*.jpg *.jpeg"),
                ("PNG", "*.png"),
                ("Todos los archivos", "*.*")
            ]
        )

        if self.foto_path:
            # Extraer fecha y hora de los metadatos EXIF
            self.metadata_foto = self.extraer_fecha_hora_exif(self.foto_path)
            
            # Mostrar información de los metadatos
            if hasattr(self, 'info_metadata_label'):
                if self.metadata_foto['fuente'] == 'EXIF':
                    texto_info = f"📅 Foto tomada: {self.metadata_foto['fecha']} a las {self.metadata_foto['hora']}"
                    color_info = "green"
                else:
                    texto_info = f"📅 Usando fecha actual: {self.metadata_foto['fecha']} a las {self.metadata_foto['hora']}"
                    color_info = "orange"
                
                self.info_metadata_label.configure(text=texto_info, foreground=color_info)
                self.info_metadata_label.pack(pady=(5, 0))
            
            # Mostrar miniatura de la imagen
            try:
                imagen = Image.open(self.foto_path)
                imagen.thumbnail((300, 300))
                photo = ImageTk.PhotoImage(imagen)

                self.imagen_label.configure(image=photo, text="")
                self.imagen_label.image = photo  # Mantener referencia

                # Habilitar botón de análisis
                if self.ai_client:
                    self.btn_analizar.configure(state="normal")

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo cargar la imagen: {str(e)}")
                self.metadata_foto = None

    def analizar_foto(self):
        """Analizar foto con IA de Abacus.AI"""
        if not self.foto_path:
            messagebox.showwarning("Advertencia", "Por favor selecciona una foto primero")
            return

        if not self.ai_client:
            messagebox.showerror("Error", "Cliente de Abacus.AI no disponible")
            return

        # Mostrar progress bar
        self.progress.start()
        self.btn_analizar.configure(state="disabled", text="🔄 Analizando...")
        self.root.update()

        try:
            # Llamar a la API de Abacus.AI
            self.alimentos_detectados = self.ai_client.identificar_alimentos(self.foto_path)

            # Mostrar alimentos en la lista
            self.alimentos_listbox.delete(0, tk.END)
            for i, alimento in enumerate(self.alimentos_detectados, 1):
                self.alimentos_listbox.insert(tk.END, f"{i}. {alimento}")

            messagebox.showinfo("✅ Análisis Completado", 
                               f"Se detectaron {len(self.alimentos_detectados)} alimentos")

        except Exception as e:
            messagebox.showerror("❌ Error de IA", 
                               f"Error al analizar la imagen:\n{str(e)}")
            self.alimentos_detectados = []

        finally:
            # Ocultar progress bar y restaurar botón
            self.progress.stop()
            self.btn_analizar.configure(state="normal", text="🤖 Analizar con IA")

    def guardar_registro(self):
        """Guardar registro completo"""
        nombre_comida = self.nombre_comida_var.get().strip()
        if not nombre_comida:
            messagebox.showwarning("⚠️ Advertencia", "Por favor ingresa el nombre de la comida")
            return

        azucar_antes = self.azucar_antes_var.get().strip()
        azucar_despues = self.azucar_despues_var.get().strip()
        
        # Validar que al menos uno de los campos tenga valor
        if not azucar_antes and not azucar_despues:
            messagebox.showwarning("⚠️ Advertencia", "Por favor ingresa al menos un nivel de azúcar (antes o después)")
            return

        if not self.alimentos_detectados:
            messagebox.showwarning("⚠️ Advertencia", "Por favor analiza una foto primero")
            return

        # Validar y convertir valores de azúcar
        azucar_antes_valor = None
        azucar_despues_valor = None
        
        try:
            if azucar_antes:
                azucar_antes_valor = float(azucar_antes)
                if azucar_antes_valor < 0 or azucar_antes_valor > 1000:
                    messagebox.showwarning("⚠️ Advertencia", "El nivel de azúcar antes debe estar entre 0 y 1000 mg/dL")
                    return
            
            if azucar_despues:
                azucar_despues_valor = float(azucar_despues)
                if azucar_despues_valor < 0 or azucar_despues_valor > 1000:
                    messagebox.showwarning("⚠️ Advertencia", "El nivel de azúcar después debe estar entre 0 y 1000 mg/dL")
                    return
        except ValueError:
            messagebox.showerror("❌ Error", "Los niveles de azúcar deben ser números válidos")
            return

        # Crear registro usando metadatos de la foto si están disponibles
        if hasattr(self, 'metadata_foto') and self.metadata_foto:
            # Usar fecha/hora de los metadatos EXIF
            fecha_registro = self.metadata_foto['fecha']
            hora_registro = self.metadata_foto['hora']
            timestamp_registro = self.metadata_foto['datetime'].isoformat()
            fuente_fecha = self.metadata_foto['fuente']
        else:
            # Usar fecha/hora actual como fallback
            ahora = datetime.now()
            fecha_registro = ahora.strftime("%Y-%m-%d")
            hora_registro = ahora.strftime("%H:%M")
            timestamp_registro = ahora.isoformat()
            fuente_fecha = 'Actual'

        registro = {
            "fecha": fecha_registro,
            "hora": hora_registro,
            "nombre_comida": nombre_comida,
            "azucar_antes": azucar_antes_valor,
            "azucar_despues": azucar_despues_valor,
            "alimentos": self.alimentos_detectados.copy(),
            "foto_path": self.foto_path if hasattr(self, 'foto_path') else None,
            "timestamp": timestamp_registro,
            "fuente_fecha": fuente_fecha  # Nuevo campo para indicar si viene de EXIF o es actual
        }

        self.datos["registros"].append(registro)
        self.guardar_datos()

        # Crear mensaje de éxito con información de azúcar
        mensaje_azucar = ""
        if azucar_antes_valor is not None and azucar_despues_valor is not None:
            diferencia = azucar_despues_valor - azucar_antes_valor
            emoji_tendencia = "📈" if diferencia > 0 else "📉" if diferencia < 0 else "➡️"
            mensaje_azucar = f"📉 Antes: {azucar_antes_valor} mg/dL\n📈 Después: {azucar_despues_valor} mg/dL\n{emoji_tendencia} Diferencia: {diferencia:+.1f} mg/dL"
        elif azucar_antes_valor is not None:
            mensaje_azucar = f"📉 Antes de comer: {azucar_antes_valor} mg/dL"
        else:
            mensaje_azucar = f"📈 Después de comer: {azucar_despues_valor} mg/dL"

        # Mensaje de éxito con información sobre la fuente de la fecha
        if fuente_fecha == 'EXIF':
            mensaje_fecha = f"📅 Fecha de la foto: {fecha_registro} a las {hora_registro}"
        else:
            mensaje_fecha = f"📅 Fecha actual: {fecha_registro} a las {hora_registro}"

        messagebox.showinfo("✅ Éxito", 
                           f"Registro guardado: {nombre_comida}\n"
                           f"{mensaje_azucar}\n"
                           f"🍽️ Alimentos detectados: {len(self.alimentos_detectados)}\n"
                           f"{mensaje_fecha}")

        # Limpiar formulario
        self.limpiar_formulario()

    def mostrar_sugerencias_comida(self):
        """Mostrar ventana con sugerencias de nombres de comida"""
        sugerencias_window = tk.Toplevel(self.root)
        sugerencias_window.title("💡 Sugerencias de Comidas")
        sugerencias_window.geometry("400x500")
        sugerencias_window.resizable(False, False)

        main_frame = ttk.Frame(sugerencias_window, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="💡 Selecciona o inspírate", 
                 font=("Arial", 14, "bold")).pack(pady=(0, 15))

        # Sugerencias predeterminadas
        sugerencias_predeterminadas = [
            "🌅 Desayuno", "🍎 Media Mañana", "🍽️ Almuerzo", "🥨 Merienda", 
            "🌙 Cena", "🍪 Postre", "🥤 Bebida", "🍌 Snack Saludable",
            "🥗 Ensalada", "🍕 Comida", "🎂 Celebración", "💊 Con Medicamento"
        ]

        # Frame para sugerencias predeterminadas
        pred_frame = ttk.LabelFrame(main_frame, text="Sugerencias Comunes", padding="10")
        pred_frame.pack(fill=tk.X, pady=(0, 15))

        for i, sugerencia in enumerate(sugerencias_predeterminadas):
            if i % 2 == 0:
                fila_frame = ttk.Frame(pred_frame)
                fila_frame.pack(fill=tk.X, pady=2)
            
            ttk.Button(fila_frame, text=sugerencia, width=18,
                      command=lambda s=sugerencia: self.seleccionar_sugerencia(s, sugerencias_window)).pack(side=tk.LEFT, padx=(0, 5))

        # Frame para historial
        historial_frame = ttk.LabelFrame(main_frame, text="Tu Historial Reciente", padding="10")
        historial_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        # Obtener nombres únicos del historial
        nombres_historial = set()
        for registro in self.datos["registros"]:
            nombre = registro.get("nombre_comida")
            if nombre:
                nombres_historial.add(nombre)

        if nombres_historial:
            historial_lista = sorted(list(nombres_historial))[-10:]  # Últimos 10 únicos
            
            canvas = tk.Canvas(historial_frame, height=150)
            scrollbar = ttk.Scrollbar(historial_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )

            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            for nombre in historial_lista:
                ttk.Button(scrollable_frame, text=f"📋 {nombre}", width=35,
                          command=lambda n=nombre: self.seleccionar_sugerencia(n, sugerencias_window)).pack(pady=2, fill=tk.X)

            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
        else:
            ttk.Label(historial_frame, text="Aún no tienes historial de comidas", 
                     foreground="gray").pack()

        # Botón cerrar
        ttk.Button(main_frame, text="❌ Cerrar", 
                  command=sugerencias_window.destroy).pack(pady=(10, 0))

    def seleccionar_sugerencia(self, sugerencia, window):
        """Seleccionar una sugerencia y cerrar ventana"""
        # Limpiar emojis y texto extra
        nombre_limpio = sugerencia.split(" ", 1)[-1] if " " in sugerencia else sugerencia
        self.nombre_comida_var.set(nombre_limpio)
        window.destroy()

    def limpiar_formulario(self):
        """Limpiar todos los campos del formulario"""
        self.nombre_comida_var.set("")
        self.azucar_antes_var.set("")
        self.azucar_despues_var.set("")
        self.alimentos_listbox.delete(0, tk.END)
        self.imagen_label.configure(image="", text="📷 No hay imagen seleccionada")
        self.imagen_label.image = None
        self.foto_path = None
        self.alimentos_detectados = []
        self.btn_analizar.configure(state="disabled")
        
        # Limpiar información de metadatos
        if hasattr(self, 'info_metadata_label'):
            self.info_metadata_label.pack_forget()
            self.info_metadata_label.configure(text="")
        
        # Limpiar metadatos de foto
        if hasattr(self, 'metadata_foto'):
            self.metadata_foto = None

    def mostrar_historial(self):
        """Mostrar ventana con historial de registros agrupados por días"""
        historial_window = tk.Toplevel(self.root)
        historial_window.title("📋 Historial de Registros")
        historial_window.geometry("1400x700")

        # Frame principal
        main_frame = ttk.Frame(historial_window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Título
        titulo_frame = ttk.Frame(main_frame)
        titulo_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(titulo_frame, text="📊 Historial con Checkboxes", 
                 font=("Arial", 14, "bold")).pack(side=tk.LEFT)
        
        # Estadísticas rápidas
        total_registros = len(self.datos["registros"])
        if total_registros > 0:
            fechas_unicas = len(set(r['fecha'] for r in self.datos["registros"]))
            promedio_por_dia = total_registros / fechas_unicas if fechas_unicas > 0 else 0
            ttk.Label(titulo_frame, 
                     text=f"📈 {total_registros} registros • {fechas_unicas} días • {promedio_por_dia:.1f} reg/día",
                     font=("Arial", 10, "italic"), foreground="gray").pack(side=tk.RIGHT)

        # Frame contenedor horizontal
        contenedor_horizontal = ttk.Frame(main_frame)
        contenedor_horizontal.pack(fill=tk.BOTH, expand=True)

        # Frame para el tree (lado izquierdo)
        tree_frame = ttk.Frame(contenedor_horizontal)
        tree_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 15))

        # Crear Treeview con checkboxes
        tree = ttk.Treeview(tree_frame, show="tree headings", height=25, selectmode="none")
        
        # Configurar columnas con checkbox
        tree["columns"] = ("checkbox", "Hora", "Azúcar", "Alimentos", "Fuente")
        tree.heading("#0", text="📅 Fecha / 🍽️ Registros", anchor="w")
        tree.heading("checkbox", text="☑️", anchor="center")
        tree.heading("Hora", text="🕐 Hora")
        tree.heading("Azúcar", text="📊 Azúcar")
        tree.heading("Alimentos", text="🥗 Alimentos")
        tree.heading("Fuente", text="📷 Origen")

        # Configurar anchos de columnas
        tree.column("#0", width=200, minwidth=150)
        tree.column("checkbox", width=50, minwidth=50, anchor="center")
        tree.column("Hora", width=80, minwidth=60)
        tree.column("Azúcar", width=120, minwidth=80)
        tree.column("Alimentos", width=280, minwidth=200)
        tree.column("Fuente", width=100, minwidth=80)

        # Agrupar registros por fecha
        registros_por_fecha = {}
        for registro in self.datos["registros"]:
            fecha = registro["fecha"]
            if fecha not in registros_por_fecha:
                registros_por_fecha[fecha] = []
            registros_por_fecha[fecha].append(registro)

        # Ordenar fechas (más recientes primero)
        fechas_ordenadas = sorted(registros_por_fecha.keys(), reverse=True)

        # Agregar datos agrupados al tree
        for fecha in fechas_ordenadas:
            registros_del_dia = registros_por_fecha[fecha]
            
            # Ordenar registros del día por hora
            registros_del_dia.sort(key=lambda x: x["hora"])
            
            # Calcular estadísticas del día
            niveles_azucar_antes = [r.get("azucar_antes") for r in registros_del_dia if r.get("azucar_antes") is not None]
            niveles_azucar_despues = [r.get("azucar_despues") for r in registros_del_dia if r.get("azucar_despues") is not None]
            
            # Para compatibilidad con registros antiguos
            niveles_azucar_legacy = [r.get("nivel_azucar") for r in registros_del_dia if r.get("nivel_azucar") is not None]
            
            # Combinar todos los niveles para estadísticas generales
            todos_los_niveles = niveles_azucar_antes + niveles_azucar_despues + niveles_azucar_legacy
            
            if todos_los_niveles:
                promedio_dia = sum(todos_los_niveles) / len(todos_los_niveles)
                min_dia = min(todos_los_niveles)
                max_dia = max(todos_los_niveles)
            else:
                promedio_dia = 0
                min_dia = 0
                max_dia = 0
            
            # Formatear fecha para mostrar
            try:
                fecha_obj = datetime.strptime(fecha, "%Y-%m-%d")
                fecha_formateada = fecha_obj.strftime("%A, %d de %B de %Y")
                # Traducir días de la semana
                dias_semana = {
                    'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
                    'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
                }
                for eng, esp in dias_semana.items():
                    fecha_formateada = fecha_formateada.replace(eng, esp)
                
                # Traducir meses
                meses = {
                    'January': 'enero', 'February': 'febrero', 'March': 'marzo', 'April': 'abril',
                    'May': 'mayo', 'June': 'junio', 'July': 'julio', 'August': 'agosto',
                    'September': 'septiembre', 'October': 'octubre', 'November': 'noviembre', 'December': 'diciembre'
                }
                for eng, esp in meses.items():
                    fecha_formateada = fecha_formateada.replace(eng, esp)
                    
            except:
                fecha_formateada = fecha
            
            # Crear nodo padre para el día
            texto_dia = f"📅 {fecha_formateada}"
            estadisticas_dia = f"📊 Promedio: {promedio_dia:.0f} mg/dL (↓{min_dia} ↑{max_dia}) • {len(registros_del_dia)} registros"
            
            dia_id = tree.insert("", "end", text=texto_dia, values=("", "", estadisticas_dia, "", "", ""), open=True)
            
            # Inicializar mapeo si es necesario
            if not hasattr(self, '_tree_registro_map'):
                self._tree_registro_map = {}
            if not hasattr(self, '_checkbox_states'):
                self._checkbox_states = {}
            
            # Agregar registros como hijos con checkboxes
            for registro in registros_del_dia:
                comida_nombre = registro.get("nombre_comida", registro.get("tipo_comida", "Sin nombre"))
                alimentos_str = ", ".join(registro["alimentos"][:3])  # Mostrar solo primeros 3 alimentos
                if len(registro["alimentos"]) > 3:
                    alimentos_str += f" (+{len(registro['alimentos'])-3} más)"
                
                # Mostrar fuente de la fecha
                fuente_fecha = registro.get("fuente_fecha", "Manual")
                icono_fuente = "📷 EXIF" if fuente_fecha == "EXIF" else "⏰ Manual"
                
                # Construir texto de azúcar
                azucar_antes = registro.get("azucar_antes")
                azucar_despues = registro.get("azucar_despues")
                nivel_azucar_legacy = registro.get("nivel_azucar")  # Para compatibilidad
                
                if azucar_antes is not None and azucar_despues is not None:
                    # Ambos valores disponibles
                    diferencia = azucar_despues - azucar_antes
                    emoji_tendencia = "📈" if diferencia > 0 else "📉" if diferencia < 0 else "➡️"
                    azucar_texto = f"📉{azucar_antes} → 📈{azucar_despues} {emoji_tendencia}"
                elif azucar_antes is not None:
                    # Solo antes
                    azucar_texto = f"📉 {azucar_antes} mg/dL (antes)"
                elif azucar_despues is not None:
                    # Solo después
                    azucar_texto = f"📈 {azucar_despues} mg/dL (después)"
                elif nivel_azucar_legacy is not None:
                    # Registro antiguo
                    nivel = nivel_azucar_legacy
                    if nivel < 70:
                        emoji_nivel = "🔵"
                    elif nivel <= 140:
                        emoji_nivel = "✅"
                    elif nivel <= 200:
                        emoji_nivel = "⚠️"
                    else:
                        emoji_nivel = "🔴"
                    azucar_texto = f"{emoji_nivel} {nivel} mg/dL"
                else:
                    azucar_texto = "Sin datos"
                
                item_id = tree.insert(dia_id, "end", 
                           text=f"   🍽️ {comida_nombre}",
                           values=(
                               "☐",  # Checkbox inicial (sin marcar)
                               registro["hora"],
                               azucar_texto,
                               alimentos_str,
                               icono_fuente
                           ))
                
                # Mapear item_id al registro para poder acceder después
                self._tree_registro_map[item_id] = registro
                self._checkbox_states[item_id] = False  # Estado inicial: sin marcar

        # Función para manejar clicks en checkboxes
        def on_tree_click(event):
            item = tree.identify_row(event.y)
            if item and item in self._tree_registro_map:  # Solo registros, no fechas
                column = tree.identify_column(event.x)
                if column == "#1":  # Columna de checkbox
                    # Cambiar estado
                    current_state = self._checkbox_states.get(item, False)
                    new_state = not current_state
                    self._checkbox_states[item] = new_state
                    
                    # Actualizar visual
                    checkbox_text = "☑️" if new_state else "☐"
                    values = list(tree.item(item, "values"))
                    values[0] = checkbox_text
                    tree.item(item, values=values)
        
        tree.bind("<Button-1>", on_tree_click)

        # Scrollbars (modificadas para el nuevo layout)
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Layout con grid para mejor control
        tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Panel de botones (lado derecho, vertical)
        botones_frame = ttk.Frame(contenedor_horizontal)
        botones_frame.pack(side=tk.RIGHT, fill=tk.Y)
        
        ttk.Label(botones_frame, text="📁 Opciones", font=("Arial", 12, "bold")).pack(pady=(0, 10))
        ttk.Label(botones_frame, text="💡 Marca los registros\ncon ☑️ para usarlos", 
                 font=("Arial", 9, "italic"), foreground="gray", justify=tk.CENTER).pack(pady=(0, 15))
        
        # Botones principales (verticales)
        ttk.Button(botones_frame, text="📊 Exportar a Excel", width=18,
                  command=lambda: self.exportar_excel_checkboxes()).pack(pady=(0, 8))
        ttk.Button(botones_frame, text="📄 Exportar a CSV", width=18,
                  command=lambda: self.exportar_csv_checkboxes()).pack(pady=(0, 8))
        ttk.Button(botones_frame, text="🗑️ Borrar registros", width=18,
                  command=lambda: self.borrar_checkboxes(historial_window)).pack(pady=(0, 8))
        
        # Separador
        ttk.Separator(botones_frame, orient='horizontal').pack(fill=tk.X, pady=15)
        
        # Botones de selección rápida
        ttk.Label(botones_frame, text="Selección rápida:", font=("Arial", 10, "bold")).pack(pady=(0, 8))
        ttk.Button(botones_frame, text="✅ Marcar todo", width=18,
                  command=lambda: self.marcar_todos_checkboxes(True)).pack(pady=(0, 4))
        ttk.Button(botones_frame, text="❌ Desmarcar todo", width=18,
                  command=lambda: self.marcar_todos_checkboxes(False)).pack(pady=(0, 15))
        
        ttk.Button(botones_frame, text="❌ Cerrar", width=18,
                  command=historial_window.destroy).pack(pady=(10, 0))

        # Guardar referencia al tree para los métodos
        self.current_tree = tree

    def configurar_horarios(self):
        """Ventana para configurar franjas horarias con nombres personalizables"""
        config_window = tk.Toplevel(self.root)
        config_window.title("⚙️ Configurar Horarios")
        config_window.geometry("500x400")

        main_frame = ttk.Frame(config_window, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="🕐 Configurar Franjas Horarias", 
                 font=("Arial", 14, "bold")).pack(pady=(0, 15))

        # Frame con scroll para los horarios
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Variables para horarios y nombres
        self.horarios_vars = {}
        self.nombres_vars = {}

        # Encabezados
        headers_frame = ttk.Frame(scrollable_frame)
        headers_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(headers_frame, text="Nombre de Comida", width=18, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Label(headers_frame, text="Hora Inicio", width=10, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Label(headers_frame, text="Hora Fin", width=10, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Label(headers_frame, text="Acciones", width=10, font=("Arial", 10, "bold")).pack(side=tk.LEFT)

        for comida, franja in self.datos["configuracion"]["franjas_horarias"].items():
            frame = ttk.Frame(scrollable_frame)
            frame.pack(fill=tk.X, pady=3)

            # Campo para el nombre de la comida
            nombre_var = tk.StringVar(value=comida)
            nombre_entry = ttk.Entry(frame, textvariable=nombre_var, width=18)
            nombre_entry.pack(side=tk.LEFT, padx=(0, 10))

            # Campos para horarios
            inicio_var = tk.StringVar(value=franja["inicio"])
            fin_var = tk.StringVar(value=franja["fin"])

            ttk.Entry(frame, textvariable=inicio_var, width=8).pack(side=tk.LEFT, padx=(0, 5))
            ttk.Label(frame, text="a").pack(side=tk.LEFT, padx=2)
            ttk.Entry(frame, textvariable=fin_var, width=8).pack(side=tk.LEFT, padx=(5, 10))

            # Botón para eliminar
            ttk.Button(frame, text="🗑️", width=3,
                      command=lambda c=comida: self.eliminar_horario(c, scrollable_frame)).pack(side=tk.LEFT)

            self.horarios_vars[comida] = {"inicio": inicio_var, "fin": fin_var}
            self.nombres_vars[comida] = nombre_var

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Botón para agregar nueva comida
        ttk.Button(main_frame, text="➕ Agregar Nueva Comida", 
                  command=lambda: self.agregar_nueva_comida(scrollable_frame)).pack(pady=(10, 0))

        # Botones principales
        botones_frame = ttk.Frame(main_frame)
        botones_frame.pack(pady=20)

        ttk.Button(botones_frame, text="💾 Guardar", 
                  command=lambda: self.guardar_horarios(config_window)).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(botones_frame, text="↺ Restaurar Predeterminados", 
                  command=lambda: self.restaurar_horarios_predeterminados(config_window)).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(botones_frame, text="❌ Cancelar", 
                  command=config_window.destroy).pack(side=tk.LEFT)

    def agregar_nueva_comida(self, parent_frame):
        """Agregar una nueva comida personalizada"""
        # Crear frame para nueva comida
        frame = ttk.Frame(parent_frame)
        frame.pack(fill=tk.X, pady=3)

        # Campo para el nombre (valor por defecto)
        nombre_var = tk.StringVar(value="Nueva Comida")
        nombre_entry = ttk.Entry(frame, textvariable=nombre_var, width=18)
        nombre_entry.pack(side=tk.LEFT, padx=(0, 10))

        # Campos para horarios (valores por defecto)
        inicio_var = tk.StringVar(value="12:00")
        fin_var = tk.StringVar(value="13:00")

        ttk.Entry(frame, textvariable=inicio_var, width=8).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Label(frame, text="a").pack(side=tk.LEFT, padx=2)
        ttk.Entry(frame, textvariable=fin_var, width=8).pack(side=tk.LEFT, padx=(5, 10))

        # Botón para eliminar
        comida_key = f"nueva_comida_{len(self.horarios_vars)}"
        ttk.Button(frame, text="🗑️", width=3,
                  command=lambda: self.eliminar_horario_temporal(frame)).pack(side=tk.LEFT)

        # Guardar variables
        self.horarios_vars[comida_key] = {"inicio": inicio_var, "fin": fin_var}
        self.nombres_vars[comida_key] = nombre_var

        # Hacer scroll hacia abajo
        parent_frame.update_idletasks()

    def eliminar_horario(self, comida_key, parent_frame):
        """Eliminar un horario existente"""
        if len(self.horarios_vars) <= 1:
            messagebox.showwarning("⚠️ Advertencia", "Debe mantener al menos una comida configurada")
            return

        # Confirmar eliminación
        resultado = messagebox.askyesno("🗑️ Confirmar Eliminación", 
                                      f"¿Está seguro de eliminar '{self.nombres_vars[comida_key].get()}'?")
        if resultado:
            # Eliminar de las variables
            del self.horarios_vars[comida_key]
            del self.nombres_vars[comida_key]
            
            # Actualizar la interfaz
            messagebox.showinfo("✅ Eliminado", "Comida eliminada correctamente")
            parent_frame.master.master.destroy()  # Cerrar ventana
            self.configurar_horarios()  # Reabrir ventana actualizada

    def eliminar_horario_temporal(self, frame):
        """Eliminar un horario temporal (recién agregado)"""
        frame.destroy()

    def restaurar_horarios_predeterminados(self, window):
        """Restaurar horarios a valores predeterminados"""
        resultado = messagebox.askyesno("↺ Restaurar Predeterminados", 
                                      "¿Está seguro de restaurar todos los horarios a los valores predeterminados?\n\nEsto eliminará todas las comidas personalizadas.")
        if resultado:
            self.datos["configuracion"]["franjas_horarias"] = {
                "desayuno": {"inicio": "06:00", "fin": "10:00"},
                "almuerzo": {"inicio": "10:01", "fin": "13:00"},
                "comida": {"inicio": "13:01", "fin": "16:00"},
                "merienda": {"inicio": "16:01", "fin": "19:00"},
                "cena": {"inicio": "19:01", "fin": "23:59"}
            }
            self.guardar_datos()
            messagebox.showinfo("✅ Restaurado", "Horarios restaurados a valores predeterminados")
            window.destroy()

    def guardar_horarios(self, window):
        """Guardar configuración de horarios con nombres personalizados"""
        try:
            # Crear nueva configuración
            nueva_configuracion = {}
            nombres_usados = set()

            for comida_key in list(self.horarios_vars.keys()):
                nuevo_nombre = self.nombres_vars[comida_key].get().strip().lower()
                inicio = self.horarios_vars[comida_key]["inicio"].get()
                fin = self.horarios_vars[comida_key]["fin"].get()

                # Validaciones
                if not nuevo_nombre:
                    messagebox.showerror("❌ Error", "El nombre de la comida no puede estar vacío")
                    return

                if nuevo_nombre in nombres_usados:
                    messagebox.showerror("❌ Error", f"El nombre '{nuevo_nombre}' ya está en uso")
                    return

                # Validar formato de hora
                try:
                    datetime.strptime(inicio, "%H:%M")
                    datetime.strptime(fin, "%H:%M")
                except ValueError:
                    messagebox.showerror("❌ Error", f"Formato de hora inválido en '{nuevo_nombre}'. Use HH:MM")
                    return

                # Validar que hora inicio sea menor que hora fin
                if datetime.strptime(inicio, "%H:%M").time() >= datetime.strptime(fin, "%H:%M").time():
                    messagebox.showerror("❌ Error", f"La hora de inicio debe ser menor que la hora de fin en '{nuevo_nombre}'")
                    return

                nombres_usados.add(nuevo_nombre)
                nueva_configuracion[nuevo_nombre] = {
                    "inicio": inicio,
                    "fin": fin
                }

            # Verificar que no haya solapamientos de horarios
            horarios_lista = []
            for nombre, horario in nueva_configuracion.items():
                inicio = datetime.strptime(horario["inicio"], "%H:%M").time()
                fin = datetime.strptime(horario["fin"], "%H:%M").time()
                horarios_lista.append((inicio, fin, nombre))

            horarios_lista.sort()
            for i in range(len(horarios_lista) - 1):
                if horarios_lista[i][1] >= horarios_lista[i + 1][0]:
                    messagebox.showwarning("⚠️ Advertencia", 
                                         f"Hay solapamiento entre '{horarios_lista[i][2]}' y '{horarios_lista[i + 1][2]}'.\n\nSe guardará de todas formas, pero podría causar confusión.")

            # Guardar configuración
            self.datos["configuracion"]["franjas_horarias"] = nueva_configuracion
            self.guardar_datos()
            messagebox.showinfo("✅ Éxito", "Horarios y nombres guardados correctamente")
            window.destroy()

        except Exception as e:
            messagebox.showerror("❌ Error", f"Error al guardar: {str(e)}")

    def exportar_csv(self):
        """Exportar historial a CSV"""
        if not self.datos["registros"]:
            messagebox.showwarning("⚠️ Advertencia", "No hay registros para exportar")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="Exportar historial a CSV"
        )

        if filename:
            try:
                import csv
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    fieldnames = ['Fecha', 'Hora', 'Nombre_Comida', 'Azucar_Antes_mg/dL', 'Azucar_Despues_mg/dL', 'Nivel_Azucar_Legacy_mg/dL', 'Alimentos_Detectados']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                    writer.writeheader()
                    for registro in self.datos["registros"]:
                        # Usar nombre_comida si existe, sino usar tipo_comida para compatibilidad
                        nombre_comida = registro.get("nombre_comida", registro.get("tipo_comida", "Sin nombre"))
                        writer.writerow({
                            'Fecha': registro['fecha'],
                            'Hora': registro['hora'],
                            'Nombre_Comida': nombre_comida,
                            'Azucar_Antes_mg/dL': registro.get('azucar_antes', ''),
                            'Azucar_Despues_mg/dL': registro.get('azucar_despues', ''),
                            'Nivel_Azucar_Legacy_mg/dL': registro.get('nivel_azucar', ''),
                            'Alimentos_Detectados': ', '.join(registro['alimentos'])
                        })

                messagebox.showinfo("✅ Éxito", f"📄 Datos exportados correctamente a:\n{filename}")
            except Exception as e:
                messagebox.showerror("❌ Error", f"Error al exportar: {str(e)}")

    def marcar_todos_checkboxes(self, marcar=True):
        """Marcar o desmarcar todos los checkboxes"""
        if not hasattr(self, '_checkbox_states') or not hasattr(self, 'current_tree'):
            return
            
        tree = self.current_tree
        checkbox_text = "☑️" if marcar else "☐"
        
        for item_id in self._checkbox_states:
            self._checkbox_states[item_id] = marcar
            values = list(tree.item(item_id, "values"))
            if values:
                values[0] = checkbox_text
                tree.item(item_id, values=values)

    def obtener_registros_marcados(self):
        """Obtener lista de registros con checkboxes marcados"""
        if not hasattr(self, '_checkbox_states') or not hasattr(self, '_tree_registro_map'):
            return []
            
        registros_marcados = []
        for item_id, marcado in self._checkbox_states.items():
            if marcado and item_id in self._tree_registro_map:
                registros_marcados.append(self._tree_registro_map[item_id])
                
        return registros_marcados

    def exportar_excel_checkboxes(self):
        """Exportar registros marcados a Excel"""
        registros = self.obtener_registros_marcados()
        
        if not registros:
            messagebox.showwarning("⚠️ Sin selección", 
                                 "Marca al menos un registro (☑️) para exportar.\n\n"
                                 "💡 Haz click en la columna ☑️ para marcar registros.")
            return
            
        # Exportar solo los registros marcados
        self.exportar_excel_filtrado(registros)

    def exportar_csv_checkboxes(self):
        """Exportar registros marcados a CSV"""
        registros = self.obtener_registros_marcados()
        
        if not registros:
            messagebox.showwarning("⚠️ Sin selección", 
                                 "Marca al menos un registro (☑️) para exportar.\n\n"
                                 "💡 Haz click en la columna ☑️ para marcar registros.")
            return
            
        # Exportar solo los registros marcados
        self.exportar_csv_filtrado(registros)

    def borrar_checkboxes(self, ventana):
        """Borrar registros marcados"""
        registros = self.obtener_registros_marcados()
        
        if not registros:
            messagebox.showwarning("⚠️ Sin selección", 
                                 "Marca al menos un registro (☑️) para borrar.\n\n"
                                 "💡 Haz click en la columna ☑️ para marcar registros.")
            return
            
        if not messagebox.askyesno("Confirmar borrado", 
                                  f"¿Seguro que deseas borrar {len(registros)} registro(s)?\n\n"
                                  "Esta acción no se puede deshacer."):
            return
            
        # Eliminar registros
        self.datos["registros"] = [r for r in self.datos["registros"] if r not in registros]
        self.guardar_datos()
        messagebox.showinfo("✅ Borrado", f"Se han borrado {len(registros)} registro(s)")
        ventana.destroy()
        self.mostrar_historial()

    def exportar_excel_filtrado(self, registros_filtrados):
        """Exportar lista específica de registros a Excel"""
        if not registros_filtrados:
            messagebox.showwarning("⚠️ Advertencia", "No hay registros para exportar")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Exportar registros marcados a Excel"
        )

        if filename:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                from datetime import datetime as dt

                # Crear libro
                wb = Workbook()
                ws = wb.active
                ws.title = "Registros Marcados"

                # Estilos
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

                # Título
                ws['A1'] = "📊 REGISTROS SELECCIONADOS"
                ws['A1'].font = Font(bold=True, size=16, color="4472C4")
                ws.merge_cells('A1:F1')
                ws['A1'].alignment = Alignment(horizontal='center')

                # Info
                ws['A2'] = f"📅 Generado: {dt.now().strftime('%d/%m/%Y %H:%M')}"
                ws['A3'] = f"📊 Registros: {len(registros_filtrados)}"

                # Encabezados
                headers = ['📅 Fecha', '🕐 Hora', '🍽️ Comida', '📉 Azúcar Antes', '📈 Azúcar Después', '🥗 Alimentos']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=5, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border

                # Datos
                for row, registro in enumerate(registros_filtrados, start=6):
                    nombre_comida = registro.get("nombre_comida", registro.get("tipo_comida", "Sin nombre"))
                    alimentos_str = ', '.join(registro['alimentos'])
                    
                    azucar_antes = registro.get('azucar_antes', '')
                    azucar_despues = registro.get('azucar_despues', '')
                    
                    azucar_antes_str = f"{azucar_antes} mg/dL" if azucar_antes else "-"
                    azucar_despues_str = f"{azucar_despues} mg/dL" if azucar_despues else "-"

                    datos_fila = [
                        registro['fecha'], registro['hora'], nombre_comida,
                        azucar_antes_str, azucar_despues_str, alimentos_str
                    ]

                    for col, valor in enumerate(datos_fila, 1):
                        cell = ws.cell(row=row, column=col, value=valor)
                        cell.border = border

                # Ajustar anchos
                for col, width in enumerate([12, 8, 20, 15, 15, 40], 1):
                    ws.column_dimensions[get_column_letter(col)].width = width

                wb.save(filename)
                messagebox.showinfo("✅ Éxito", f"📊 {len(registros_filtrados)} registros exportados a Excel")
                
            except ImportError:
                messagebox.showerror("❌ Error", "Librería openpyxl no instalada.\nEjecuta: pip install openpyxl")
            except Exception as e:
                messagebox.showerror("❌ Error", f"Error al exportar: {str(e)}")

    def exportar_csv_filtrado(self, registros_filtrados):
        """Exportar lista específica de registros a CSV"""
        if not registros_filtrados:
            messagebox.showwarning("⚠️ Advertencia", "No hay registros para exportar")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="Exportar registros marcados a CSV"
        )

        if filename:
            try:
                import csv
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    
                    # Encabezados
                    writer.writerow(['Fecha', 'Hora', 'Comida', 'Azúcar Antes', 'Azúcar Después', 'Alimentos'])
                    
                    # Datos
                    for registro in registros_filtrados:
                        nombre_comida = registro.get("nombre_comida", registro.get("tipo_comida", "Sin nombre"))
                        alimentos_str = ', '.join(registro['alimentos'])
                        
                        azucar_antes = registro.get('azucar_antes', '')
                        azucar_despues = registro.get('azucar_despues', '')
                        
                        writer.writerow([
                            registro['fecha'], registro['hora'], nombre_comida,
                            azucar_antes if azucar_antes else '',
                            azucar_despues if azucar_despues else '',
                            alimentos_str
                        ])

                messagebox.showinfo("✅ Éxito", f"📄 {len(registros_filtrados)} registros exportados a CSV")
                
            except Exception as e:
                messagebox.showerror("❌ Error", f"Error al exportar: {str(e)}")

    def exportar_excel(self):
        """Exportar historial a Excel con formato mejorado"""
        if not self.datos["registros"]:
            messagebox.showwarning("⚠️ Advertencia", "No hay registros para exportar")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Exportar historial a Excel"
        )

        if filename:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                from datetime import datetime as dt

                # Crear libro de trabajo
                wb = Workbook()
                ws = wb.active
                ws.title = "Control de Azúcar"

                # Configurar estilos
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Título del documento
                ws['A1'] = "📊 CONTROL DE AZÚCAR Y ALIMENTACIÓN"
                ws['A1'].font = Font(bold=True, size=16, color="4472C4")
                ws.merge_cells('A1:F1')
                ws['A1'].alignment = Alignment(horizontal='center')

                # Información del reporte
                ws['A2'] = f"📅 Generado el: {dt.now().strftime('%d/%m/%Y %H:%M')}"
                ws['A2'].font = Font(size=10, italic=True)
                ws['A3'] = f"📊 Total de registros: {len(self.datos['registros'])}"
                ws['A3'].font = Font(size=10, italic=True)

                # Encabezados de la tabla (fila 5)
                headers = ['📅 Fecha', '🕐 Hora', '🍽️ Comida', '� Azúcar Antes', '�📈 Azúcar Después', '🥗 Alimentos', '📊 Estado']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=5, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                # Datos de los registros
                for row, registro in enumerate(self.datos["registros"], start=6):
                    # Usar nombre_comida si existe, sino usar tipo_comida para compatibilidad
                    nombre_comida = registro.get("nombre_comida", registro.get("tipo_comida", "Sin nombre"))
                    alimentos_str = ', '.join(registro['alimentos'])
                    
                    # Obtener valores de azúcar
                    azucar_antes = registro.get('azucar_antes', '')
                    azucar_despues = registro.get('azucar_despues', '')
                    nivel_azucar_legacy = registro.get('nivel_azucar')  # Para compatibilidad
                    
                    # Determinar estado del azúcar basado en cualquier valor disponible
                    niveles_para_estado = []
                    if azucar_antes:
                        niveles_para_estado.append(azucar_antes)
                    if azucar_despues:
                        niveles_para_estado.append(azucar_despues)
                    if nivel_azucar_legacy and not niveles_para_estado:
                        niveles_para_estado.append(nivel_azucar_legacy)
                    
                    if niveles_para_estado:
                        nivel_para_estado = max(niveles_para_estado)  # Usar el más alto para determinar estado
                        if nivel_para_estado < 70:
                            estado = "🔻 Bajo"
                            estado_color = "FF6B6B"
                        elif nivel_para_estado <= 140:
                            estado = "✅ Normal"
                            estado_color = "51CF66"
                        elif nivel_para_estado <= 200:
                            estado = "⚠️ Alto"
                            estado_color = "FFD43B"
                        else:
                            estado = "🔴 Muy Alto"
                            estado_color = "FF6B6B"
                    else:
                        estado = "❓ Sin datos"
                        estado_color = "CCCCCC"

                    # Formatear valores para mostrar
                    azucar_antes_str = f"{azucar_antes} mg/dL" if azucar_antes else "-"
                    azucar_despues_str = f"{azucar_despues} mg/dL" if azucar_despues else "-"
                    
                    # Para compatibilidad con registros antiguos
                    if nivel_azucar_legacy and not azucar_antes and not azucar_despues:
                        azucar_antes_str = f"{nivel_azucar_legacy} mg/dL (legacy)"
                        azucar_despues_str = "-"

                    # Escribir datos
                    datos_fila = [
                        registro['fecha'],
                        registro['hora'],
                        nombre_comida,
                        azucar_antes_str,
                        azucar_despues_str,
                        alimentos_str,
                        estado
                    ]

                    for col, valor in enumerate(datos_fila, 1):
                        cell = ws.cell(row=row, column=col, value=valor)
                        cell.border = border
                        cell.alignment = Alignment(vertical='center')
                        
                        # Colorear la celda de estado según el nivel de azúcar
                        if col == 7:  # Columna de estado (ahora es la 7ª columna)
                            cell.fill = PatternFill(start_color=estado_color, end_color=estado_color, fill_type="solid")
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                # Ajustar ancho de columnas
                column_widths = [12, 8, 20, 15, 15, 40, 12]
                for col, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(col)].width = width

                # Agregar hoja de estadísticas
                stats_ws = wb.create_sheet("📊 Estadísticas")
                
                # Estadísticas básicas - recopilar todos los niveles de azúcar
                todos_los_niveles = []
                for r in self.datos["registros"]:
                    if r.get('azucar_antes'):
                        todos_los_niveles.append(r['azucar_antes'])
                    if r.get('azucar_despues'):
                        todos_los_niveles.append(r['azucar_despues'])
                    if r.get('nivel_azucar') and not r.get('azucar_antes') and not r.get('azucar_despues'):
                        todos_los_niveles.append(r['nivel_azucar'])
                
                if todos_los_niveles:
                    promedio = sum(todos_los_niveles) / len(todos_los_niveles)
                    maximo = max(todos_los_niveles)
                    minimo = min(todos_los_niveles)
                    
                    normal_count = len([n for n in todos_los_niveles if 70 <= n <= 140])
                    alto_count = len([n for n in todos_los_niveles if n > 140])
                    bajo_count = len([n for n in todos_los_niveles if n < 70])

                    stats_data = [
                        ["📊 ESTADÍSTICAS DE AZÚCAR", ""],
                        ["", ""],
                        ["📈 Nivel promedio", f"{promedio:.1f} mg/dL"],
                        ["🔺 Nivel máximo", f"{maximo} mg/dL"],
                        ["🔻 Nivel mínimo", f"{minimo} mg/dL"],
                        ["", ""],
                        ["📊 DISTRIBUCIÓN", ""],
                        ["✅ Registros normales (70-140)", normal_count],
                        ["⚠️ Registros altos (>140)", alto_count],
                        ["🔻 Registros bajos (<70)", bajo_count],
                        ["", ""],
                        ["📈 Porcentaje normal", f"{(normal_count/len(todos_los_niveles)*100):.1f}%"],
                        ["⚠️ Porcentaje alto", f"{(alto_count/len(todos_los_niveles)*100):.1f}%"],
                        ["🔻 Porcentaje bajo", f"{(bajo_count/len(todos_los_niveles)*100):.1f}%"]
                    ]

                    for row, (label, value) in enumerate(stats_data, 1):
                        stats_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                        stats_ws.cell(row=row, column=2, value=value)
                    
                    # Ajustar anchos
                    stats_ws.column_dimensions['A'].width = 25
                    stats_ws.column_dimensions['B'].width = 15

                # Guardar archivo
                wb.save(filename)
                messagebox.showinfo("✅ Éxito", f"📊 Datos exportados correctamente a Excel:\n{filename}\n\nSe incluyeron:\n• Hoja principal con datos\n• Hoja de estadísticas\n• Formato con colores según niveles")
                
            except ImportError:
                messagebox.showerror("❌ Error", "La librería openpyxl no está instalada.\nEjecuta: pip install openpyxl")
            except Exception as e:
                messagebox.showerror("❌ Error", f"Error al exportar a Excel: {str(e)}")

# Función principal
def main():
    """Función principal para ejecutar la aplicación"""
    # Verificar que existe el archivo .env
    if not os.path.exists('.env'):
        messagebox.showerror("Error de Configuración", 
                           "No se encontró el archivo .env\n\n"
                           "Crea un archivo .env con tu ABACUS_API_KEY")
        return

    root = tk.Tk()
    app = ControlAzucarApp(root)

    # Configurar cierre de aplicación
    def on_closing():
        if messagebox.askokcancel("Salir", "¿Deseas cerrar la aplicación?"):
            root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing)
    root.mainloop()

if __name__ == "__main__":
    main()
